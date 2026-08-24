# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from comments import LEVELS, table_A, table_B, table_R, table_C, COMBO_DESC, rare_note

FONT = "Microsoft JhengHei"
f_body = Font(name=FONT, size=10)
f_bold = Font(name=FONT, size=10, bold=True)
f_title = Font(name=FONT, size=13, bold=True)
f_blue = Font(name=FONT, size=10, color="0000FF")
fill_head = PatternFill("solid", fgColor="D9D9D9")
fill_in = PatternFill("solid", fgColor="FFFF99")
fill_A = PatternFill("solid", fgColor="FFF2CC")
fill_B = PatternFill("solid", fgColor="E2EFDA")
fill_R = PatternFill("solid", fgColor="DDEBF7")
fill_C = PatternFill("solid", fgColor="FCE4D6")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top", wrap_text=True)

wb = Workbook()
ORDER = ["N1", "N2", "N3", "N4", "N5"]

# ================= 說明 =================
ws = wb.active; ws.title = "說明"
lines = [
    ("2026 日檢弱點分析 評語清單 N1～N5（模組整合版）", f_title),
    ("", f_body),
    ("一、組裝順序", f_bold),
    ("開頭句（表A）→ 言語知識評語（表B，N4/N5 含讀解）→ 讀解評語（表R，僅 N1～N3）→ 聽解評語（表C）→ 結尾句（表A）", f_body),
    ("學生看到的是一段連續文字，五個部分直接串接。", f_body),
    ("", f_body),
    ("二、查表鍵", f_bold),
    ("「評語清單」E 欄 = 級數|表|區間代碼|參考情報，例：N5|B_言語知識|中|BA。用 INDEX/MATCH 或 VLOOKUP 對 E 欄查即可。", f_body),
    ("表A 與表C、表R 的參考情報欄固定填「-」。", f_body),
    ("", f_body),
    ("三、區間代碼", f_bold),
    ("表A（總判定，看總分與及格線的差距）：單科未達／差20+／差1-19／過0-14／過15-39／過40+", f_body),
    ("　「單科未達」的判斷優先：總分已達及格分，但任一科低於該科門檻 → 用此列。總分未達及格分時，一律依差距分列，不論單科。", f_body),
    ("表B／表R／表C（看各科分數與門檻）：未達（<門檻）／偏低／中／高，各段上限見「級別設定」。", f_body),
    ("", f_body),
    ("四、參考情報代碼（僅表B 使用）", f_bold),
    ("AA＝單字A文法A｜AB＝單字A文法B（文法偏弱）｜BA＝單字B文法A（單字偏弱）｜BB｜C1＝其中一項C｜CC＝兩項C", f_body),
    ("C1 的評語含「〇〇」，系統代入「單字」或「文法」（哪一項是 C 就代入哪一項）。", f_body),
    ("表B 的「高」段只保留 AA／AB／BA 三列，其餘組合一律走「其他」（實務上幾乎不會出現）。", f_body),
    ("備註欄標「少見」的列：N1～N3 的言語知識＝文字語彙＋文法，參考情報理論上會和分數一致，這些組合只是備用，避免查表落空。", f_body),
    ("", f_body),
    ("五、級別設定", f_bold),
    ("及格分、各科滿分與門檻依 JLPT 官方公布；分段上限（偏低上限、中上限）是本表的設計值，可依實際分布調整，調整後表B～表C 的「分數下限／上限」欄需一併改。", f_body),
    ("N4／N5 的言語知識科目為「言語知識（文字・語彙・文法）・讀解」合併計分，讀解欄填 0 表示無此科。", f_body),
    ("", f_body),
    ("六、試算", f_bold),
    ("「試算」分頁黃色欄位可直接輸入一位學生的成績，下方公式會自動判斷代碼並組出完整評語，可用來驗證邏輯，或直接把公式搬到你的系統設定。", f_body),
]
for i, (t, f) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=t); c.font = f; c.alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 120

# ================= 級別設定 =================
ws = wb.create_sheet("級別設定")
heads = ["級數", "及格分", "下一級", "言語知識科目名", "言語知識滿分", "言語知識門檻", "言語知識偏低上限", "言語知識中上限",
         "讀解滿分", "讀解門檻", "讀解偏低上限", "讀解中上限", "聽解滿分", "聽解門檻", "聽解偏低上限", "聽解中上限"]
for j, h in enumerate(heads, 1):
    c = ws.cell(row=1, column=j, value=h); c.font = f_bold; c.fill = fill_head; c.alignment = center; c.border = border
for i, lv in enumerate(ORDER, 2):
    p = LEVELS[lv]
    row = [lv, p["pass_"], p["nxt"] or "—", p["subj"], p["lk_max"], p["lk_min"], p["lk_low"], p["lk_mid"],
           60 if p["has_r"] else 0, p.get("r_min", 0), p.get("r_low", 0), p.get("r_mid", 0),
           60, p["c_min"], p["c_low"], p["c_mid"]]
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v); c.font = f_blue if j > 1 else f_bold; c.alignment = center; c.border = border
ws.cell(row=8, column=1, value="註：及格分、滿分、門檻為 JLPT 官方規定；偏低上限與中上限為本表設計值（藍字可改）。讀解滿分 0 表示該級無獨立讀解科（N4／N5 併入言語知識）。").font = f_body
for j in range(1, 17): ws.column_dimensions[get_column_letter(j)].width = 13
ws.column_dimensions["D"].width = 20
ws.freeze_panes = "B2"

# ================= 評語清單 =================
ws = wb.create_sheet("評語清單")
heads = ["級數", "表", "區間代碼", "參考情報", "查表鍵", "條件說明", "分數下限", "分數上限", "評語／開頭句", "結尾句（表A）", "備註"]
for j, h in enumerate(heads, 1):
    c = ws.cell(row=1, column=j, value=h); c.font = f_bold; c.fill = fill_head; c.alignment = center; c.border = border

r = 2
BANDS = ["未達", "偏低", "中", "高"]
COMBOS = ["AA", "AB", "BA", "BB", "C1", "CC"]

def put(row, fill):
    global r
    for j, v in enumerate(row, 1):
        c = ws.cell(row=r, column=j, value=v); c.font = f_body; c.border = border; c.fill = fill
        c.alignment = wrap if j in (6, 9, 10) else center
    r += 1

def band_ranges(mn, low, mid, mx):
    return {"未達": (0, mn - 1), "偏低": (mn, low), "中": (low + 1, mid), "高": (mid + 1, mx)}

for lv in ORDER:
    p = LEVELS[lv]; ps = p["pass_"]
    # 表A
    a_rng = {"單科未達": (ps, 180), "差20+": (0, ps - 20), "差1-19": (ps - 19, ps - 1),
             "過0-14": (ps, ps + 14), "過15-39": (ps + 15, ps + 39), "過40+": (ps + 40, 180)}
    for code, desc, opener, closer in table_A(lv):
        lo, hi = a_rng[code]
        note = "總分達及格但任一科未達門檻，優先於其他列" if code == "單科未達" else ""
        put([lv, "A_總判定", code, "-", f"{lv}|A_總判定|{code}|-", desc, lo, hi, opener, closer, note], fill_A)
    # 表B
    B = table_B(lv); rng = band_ranges(p["lk_min"], p["lk_low"], p["lk_mid"], p["lk_max"])
    for band in BANDS:
        combos = ["AA", "AB", "BA", "其他"] if band == "高" else COMBOS
        for combo in combos:
            lo, hi = rng[band]
            desc = f"{p['subj']} {lo}～{hi}／{COMBO_DESC[combo]}"
            put([lv, "B_言語知識", band, combo, f"{lv}|B_言語知識|{band}|{combo}", desc, lo, hi,
                 B[band][combo], "", rare_note(lv, band, combo)], fill_B)
    # 表R
    if p["has_r"]:
        R = table_R(lv); rng = band_ranges(p["r_min"], p["r_low"], p["r_mid"], 60)
        for band in BANDS:
            lo, hi = rng[band]
            put([lv, "R_讀解", band, "-", f"{lv}|R_讀解|{band}|-", f"讀解 {lo}～{hi}", lo, hi, R[band], "", ""], fill_R)
    # 表C
    C = table_C(lv); rng = band_ranges(p["c_min"], p["c_low"], p["c_mid"], 60)
    for band in BANDS:
        lo, hi = rng[band]
        put([lv, "C_聽解", band, "-", f"{lv}|C_聽解|{band}|-", f"聽解 {lo}～{hi}", lo, hi, C[band], "", ""], fill_C)

widths = {"A": 7, "B": 13, "C": 11, "D": 10, "E": 26, "F": 30, "G": 9, "H": 9, "I": 70, "J": 48, "K": 22}
for k, v in widths.items(): ws.column_dimensions[k].width = v
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{r-1}"
total_rows = r - 2

# ================= 試算 =================
ws = wb.create_sheet("試算")
ws.column_dimensions["A"].width = 30
for col in "BCD": ws.column_dimensions[col].width = 60
ws["A1"] = "試算：黃色欄位輸入成績，下方自動組出評語（三個範例）"; ws["A1"].font = f_title

labels = {
    3: "級數", 4: "總分", 5: "言語知識分數（N4/N5 為言語知識・讀解）", 6: "讀解分數（僅 N1～N3）", 7: "聽解分數",
    8: "參考情報 文字・語彙（A/B/C）", 9: "參考情報 文法（A/B/C）",
    11: "及格分", 12: "有獨立讀解科", 13: "總分差距", 14: "任一科未達門檻",
    15: "總判定代碼", 16: "言語知識區間", 17: "讀解區間", 18: "聽解區間",
    19: "參考情報組合", 20: "參考情報代碼（查表用）", 21: "〇〇代入",
    23: "鍵 A", 24: "鍵 B", 25: "鍵 R", 26: "鍵 C",
    28: "開頭句", 29: "言語知識評語", 30: "讀解評語", 31: "聽解評語", 32: "結尾句",
    34: "完整評語",
}
for rr, t in labels.items():
    c = ws.cell(row=rr, column=1, value=t); c.font = f_bold if rr in (34,) else f_body; c.alignment = wrap

examples = [("N5", 84, 70, "", 14, "B", "A"), ("N2", 105, 40, 28, 37, "A", "B"), ("N1", 72, 22, 30, 20, "C", "B")]
LS = "級別設定"
for ci, ex in enumerate(examples):
    col = get_column_letter(2 + ci); L = col
    for rr, v in zip(range(3, 10), ex):
        c = ws[f"{L}{rr}"]; c.value = v if v != "" else None; c.fill = fill_in; c.font = f_blue; c.border = border; c.alignment = center
    m = f"MATCH({L}$3,{LS}!$A$2:$A$6,0)"
    def idx(colletter): return f"INDEX({LS}!${colletter}$2:${colletter}$6,{m})"
    F = {}
    F[11] = f"={idx('B')}"
    F[12] = f"={idx('I')}>0"
    F[13] = f"={L}4-{L}11"
    F[14] = f"=OR({L}5<{idx('F')},AND({L}12,N({L}6)<{idx('J')}),{L}7<{idx('N')})"
    F[15] = f'=IF(AND({L}13>=0,{L}14),"單科未達",_xlfn.IFS({L}13<=-20,"差20+",{L}13<0,"差1-19",{L}13<15,"過0-14",{L}13<40,"過15-39",TRUE,"過40+"))'
    F[16] = f'=_xlfn.IFS({L}5<{idx("F")},"未達",{L}5<={idx("G")},"偏低",{L}5<={idx("H")},"中",TRUE,"高")'
    F[17] = f'=IF({L}12,_xlfn.IFS({L}6<{idx("J")},"未達",{L}6<={idx("K")},"偏低",{L}6<={idx("L")},"中",TRUE,"高"),"-")'
    F[18] = f'=_xlfn.IFS({L}7<{idx("N")},"未達",{L}7<={idx("O")},"偏低",{L}7<={idx("P")},"中",TRUE,"高")'
    F[19] = f'=IF(AND({L}8="C",{L}9="C"),"CC",IF(OR({L}8="C",{L}9="C"),"C1",{L}8&{L}9))'
    F[20] = f'=IF({L}16="高",IF(OR({L}19="AA",{L}19="AB",{L}19="BA"),{L}19,"其他"),{L}19)'
    F[21] = f'=IF({L}8="C","單字","文法")'
    F[23] = f'={L}3&"|A_總判定|"&{L}15&"|-"'
    F[24] = f'={L}3&"|B_言語知識|"&{L}16&"|"&{L}20'
    F[25] = f'=IF({L}12,{L}3&"|R_讀解|"&{L}17&"|-","")'
    F[26] = f'={L}3&"|C_聽解|"&{L}18&"|-"'
    look = lambda key, colL: f"INDEX(評語清單!${colL}$2:${colL}${r-1},MATCH({key},評語清單!$E$2:$E${r-1},0))"
    F[28] = f"={look(L+'23','I')}"
    F[29] = f'=SUBSTITUTE({look(L+"24","I")},"〇〇",{L}21)'
    F[30] = f'=IF({L}12,{look(L+"25","I")},"")'
    F[31] = f"={look(L+'26','I')}"
    F[32] = f"={look(L+'23','J')}"
    F[34] = f"={L}28&{L}29&{L}30&{L}31&{L}32"
    for rr, f in F.items():
        c = ws[f"{L}{rr}"]; c.value = f; c.font = f_body; c.border = border
        c.alignment = wrap if rr >= 28 else center
ws.row_dimensions[34].height = 260
for rr in range(28, 33): ws.row_dimensions[rr].height = 75
ws["A36"] = "說明：黃色為輸入格；B12 的讀解分數 N4/N5 留空即可。公式只用 INDEX/MATCH/IFS/SUBSTITUTE，可直接搬到 Google Sheets。"
ws["A36"].font = f_body

out = "2026日檢弱點分析_評語清單_N1-N5.xlsx"
wb.save(out)
print("rows:", total_rows, "->", out)
